
import openpyxl 
  

wb = openpyxl.Workbook() 
  

sheet = wb.active 
sheet.cell(row = 1, column = 1).value = "Họ và tên"
sheet.cell(row = 1, column = 2).value = "Ngày sinh"
sheet.cell(row = 1, column = 3).value = "Cccd"
sheet.cell(row = 1, column = 4).value = "Số báo danh"
sheet.cell(row = 1, column = 5).value = "Nơi sinh"
sheet.cell(row = 1, column = 6).value = "Mã Schoolrank"
sheet.cell(row = 1, column = 7).value = "Top Schoolrank"
sheet.cell(row = 1, column = 8).value = "Điểm trắc nghiệm"
sheet.cell(row = 1, column = 9).value = "Điểm luận"
sheet.cell(row = 1, column = 10).value = "Điểm tổng"
sheet.cell(row = 1, column = 11).value = "Kết quả Học bổng"
wb.save("sample.xlsx")